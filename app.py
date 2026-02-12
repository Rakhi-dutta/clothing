import os
import datetime
import sqlite3
from math import ceil
from functools import wraps

from flask import (
    Flask, render_template, request, redirect,
    session, url_for, send_from_directory, flash
)

import openpyxl              # Excel
import qrcode                # QR code
from barcode import Code128  # Barcode
from barcode.writer import ImageWriter

from utils_codes import generate_barcode, generate_qr
import zipfile
from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


# ---------------- APP CONFIG ----------------
app = Flask(__name__)
app.secret_key = "supersecretkey123"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------------- DB HELPERS ----------------
def get_db():
    conn = sqlite3.connect(os.path.join(BASE_DIR, "inventory.db"))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()

    # Admins (with role)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS admin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT DEFAULT 'staff'
        )
    """)

    # default super admin
    conn.execute("""
        INSERT OR IGNORE INTO admin (id, username, password, role)
        VALUES (1, 'admin', 'admin123', 'superadmin')
    """)

    # Categories
    conn.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            description TEXT
        )
    """)

    # Clothing (main table)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clothing (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            category TEXT,
            size TEXT,
            quantity INTEGER,
            price REAL,
            created_at TEXT,
            image TEXT,
            barcode TEXT,
            qrcode TEXT
        )
    """)

    # Extra gallery images
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clothing_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clothing_id INTEGER,
            image TEXT,
            created_at TEXT
        )
    """)

    # Stock movement logs
    conn.execute("""
        CREATE TABLE IF NOT EXISTS stock_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clothing_id INTEGER,
            change_type TEXT,      -- in / out / adjust / create / import
            qty_change INTEGER,
            note TEXT,
            admin TEXT,
            created_at TEXT
        )
    """)

    # Activity logs
    conn.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT,
            details TEXT,
            created_at TEXT
        )
    """)

    # Customers
    conn.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            phone TEXT,
            address TEXT,
            city TEXT,
            state TEXT,
            zip_code TEXT,
            created_at TEXT
        )
    """)

    # Orders
    conn.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT UNIQUE NOT NULL,
            customer_id INTEGER,
            total_amount REAL,
            status TEXT DEFAULT 'pending',
            payment_status TEXT DEFAULT 'unpaid',
            shipping_address TEXT,
            notes TEXT,
            created_at TEXT,
            updated_at TEXT,
            FOREIGN KEY(customer_id) REFERENCES customers(id)
        )
    """)

    # Order Items
    conn.execute("""
        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            clothing_id INTEGER,
            size TEXT,
            quantity INTEGER,
            price REAL,
            FOREIGN KEY(order_id) REFERENCES orders(id),
            FOREIGN KEY(clothing_id) REFERENCES clothing(id)
        )
    """)

    # Notifications
    conn.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT,                 -- order_placed, order_shipped, order_delivered, payment_received
            recipient TEXT,            -- 'admin' or customer_id
            title TEXT,
            message TEXT,
            read INTEGER DEFAULT 0,
            order_id INTEGER,
            created_at TEXT
        )
    """)

    # Cart (temporary storage for users)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cart (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,
            clothing_id INTEGER,
            size TEXT,
            quantity INTEGER,
            added_at TEXT,
            FOREIGN KEY(clothing_id) REFERENCES clothing(id)
        )
    """)

    conn.commit()
    conn.close()


def log_action(action, details):
    conn = get_db()
    conn.execute(
        "INSERT INTO logs (action, details, created_at) VALUES (?, ?, ?)",
        (action, details, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()


# ---------------- AUTH HELPERS ----------------
def require_login(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def require_role(*roles):
    """Restrict route to specific roles."""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not session.get("admin"):
                return redirect(url_for("login"))
            role = session.get("role", "staff")
            if role not in roles:
                flash("You do not have permission for this action.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return wrapper
    return decorator


# ---------------- CUSTOMER/ECOM ROUTES ----------------
# Registers /shop, /cart, /checkout, and admin order/customer/notification screens.
from customer_routes import register_customer_routes

register_customer_routes(app, get_db, log_action, require_login, require_role)


# ---------------- ROOT ----------------
@app.route("/")
def root():
    return redirect(url_for("login"))


# ---------------- LOGIN ----------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form["username"]
        pw = request.form["password"]

        db = get_db()
        row = db.execute(
            "SELECT * FROM admin WHERE username=? AND password=?",
            (user, pw)
        ).fetchone()

        if row:
            session["admin"] = row["username"]
            session["role"] = row["role"]
            log_action("login", f"{row['username']} logged in")
            return redirect(url_for("dashboard"))
        else:
            return render_template("login.html", error="Wrong username or password!")

    return render_template("login.html")


# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    if session.get("admin"):
        log_action("logout", f"{session['admin']} logged out")
    session.clear()
    return redirect(url_for("login"))


# ---------------- PROFILE ----------------
@app.route("/profile")
@require_login
def profile():
    return render_template("profile.html", title="My Profile")


# ---------------- SETTINGS ----------------
@app.route("/settings")
@require_login
def settings():
    return render_template("settings.html", title="Settings")


# ---------------- CHANGE PASSWORD ----------------
@app.route("/change-password", methods=["POST"])
@require_login
def change_password():
    import json
    data = request.get_json()
    current_password = data.get("current_password")
    new_password = data.get("new_password")

    if not current_password or not new_password:
        return {"success": False, "message": "Missing password fields."}

    db = get_db()
    admin = db.execute(
        "SELECT * FROM admin WHERE username=?",
        (session.get("admin"),)
    ).fetchone()

    if not admin or admin["password"] != current_password:
        return {"success": False, "message": "Current password is incorrect."}

    if len(new_password) < 6:
        return {"success": False, "message": "New password must be at least 6 characters."}

    db.execute(
        "UPDATE admin SET password=? WHERE username=?",
        (new_password, session.get("admin"))
    )
    db.commit()

    log_action("password_change", f"{session.get('admin')} changed password")
    return {"success": True, "message": "Password changed successfully!"}


# ---------------- FORGOT PASSWORD ----------------
@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        
        if not username:
            flash("❌ Please enter your username.", "danger")
            return render_template("forgot_password.html")
        
        db = get_db()
        admin = db.execute(
            "SELECT * FROM admin WHERE username=?", (username,)
        ).fetchone()
        
        if not admin:
            flash("❌ Username not found.", "danger")
            return render_template("forgot_password.html")
        
        # In production, you would send an email here
        # For now, we'll show a recovery message
        flash(f"✅ If '{username}' exists, you will receive password recovery instructions. Contact your administrator for manual password reset.", "info")
        return render_template("forgot_password.html")
    
    return render_template("forgot_password.html", title="Forgot Password")


# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
@require_login
def dashboard():
    db = get_db()
    total_items = db.execute("SELECT COUNT(*) AS c FROM clothing").fetchone()["c"]
    total_qty = db.execute(
        "SELECT COALESCE(SUM(quantity),0) AS q FROM clothing"
    ).fetchone()["q"]
    category_count = db.execute(
        "SELECT COUNT(*) AS c FROM categories"
    ).fetchone()["c"]
    low_stock = db.execute(
        "SELECT COUNT(*) AS c FROM clothing WHERE quantity < 5"
    ).fetchone()["c"]

    recent_logs = db.execute(
        "SELECT * FROM logs ORDER BY id DESC LIMIT 5"
    ).fetchall()

    # category-wise quantity for chart
    cat_data = db.execute("""
        SELECT category, COALESCE(SUM(quantity),0) AS qty
        FROM clothing
        GROUP BY category
        ORDER BY qty DESC
        LIMIT 6
    """).fetchall()

    labels = [row["category"] for row in cat_data]
    values = [row["qty"] for row in cat_data]

    return render_template(
        "dashboard.html",
        total_items=total_items,
        total_qty=total_qty,
        category_count=category_count,
        low_stock=low_stock,
        logs=recent_logs,
        chart_labels=labels,
        chart_values=values,
        title="Dashboard"
    )


# ---------------- INVENTORY LIST ----------------
@app.route("/inventory")
@require_login
def inventory():
    db = get_db()

    # --- filters & query params ---
    q = request.args.get("q", "").strip()
    sort = request.args.get("sort", "created_desc")

    try:
        per_page = int(request.args.get("per_page", 10))
    except ValueError:
        per_page = 10

    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1
    if page < 1:
        page = 1

    # --- search filter ---
    where_clause = ""
    params = []
    if q:
        where_clause = "WHERE name LIKE ? OR category LIKE ? OR size LIKE ?"
        like = f"%{q}%"
        params.extend([like, like, like])

    # --- sorting options ---
    sort_map = {
        "name_asc": "name ASC",
        "name_desc": "name DESC",
        "qty_asc": "quantity ASC",
        "qty_desc": "quantity DESC",
        "price_asc": "price ASC",
        "price_desc": "price DESC",
        "created_desc": "created_at DESC",
        "created_asc": "created_at ASC",
    }
    order_by = sort_map.get(sort, "name ASC")

    # --- total count ---
    count_row = db.execute(
        f"SELECT COUNT(*) AS c FROM clothing {where_clause}",
        params
    ).fetchone()
    total_items = count_row["c"] if count_row else 0
    total_pages = ceil(total_items / per_page) if total_items else 1
    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    # --- actual items ---
    items = db.execute(
        f"""
        SELECT * FROM clothing
        {where_clause}
        ORDER BY {order_by}
        LIMIT ? OFFSET ?
        """,
        params + [per_page, offset]
    ).fetchall()

    pages = list(range(1, total_pages + 1))

    # 🔥🔥 HERE — add gallery images dict
    gallery_map = {}
    for it in items:
        imgs = db.execute(
            "SELECT image FROM clothing_images WHERE clothing_id=? ORDER BY id ASC",
            (it["id"],)
        ).fetchall()
        gallery_map[it["id"]] = imgs

    # --- categories ---
    cats = db.execute("SELECT * FROM categories ORDER BY name ASC").fetchall()

    return render_template(
        "inventory.html",
        items=items,
        categories=cats,
        q=q,
        sort=sort,
        per_page=per_page,
        page=page,
        pages=pages,
        total_items=total_items,

        # 🔥🔥 Pass gallery images
        gallery_map=gallery_map,

        title="Inventory"
    )


# ---------------- SEARCH ----------------
@app.route("/search")
@require_login
def search():
    q = request.args.get("q", "").strip()
    if q:
        return redirect(url_for("inventory", q=q))
    return redirect(url_for("inventory"))


# ---------------- ADD ITEM ----------------
@app.route("/inventory/add", methods=["POST"])
@require_login
def add_item():
    name = request.form["name"]
    category = request.form["category"]
    size = request.form["size"]
    quantity = int(request.form["quantity"])
    price = float(request.form["price"])

    main_image = request.files.get("image")
    gallery_files = request.files.getlist("gallery")

    db = get_db()

    # Insert item
    db.execute("""
        INSERT INTO clothing (name, category, size, quantity, price, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        name, category, size, quantity, price,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))
    db.commit()

    item_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

    # Save main image
    image_name = None
    if main_image and main_image.filename:
        image_name = f"item_{item_id}_{main_image.filename}"
        main_image.save(os.path.join(UPLOAD_FOLDER, image_name))
        db.execute("UPDATE clothing SET image=? WHERE id=?", (image_name, item_id))
        db.commit()

    # Save gallery images
    for img in gallery_files:
        if img and img.filename:
            g_name = f"gallery_{item_id}_{img.filename}"
            img.save(os.path.join(UPLOAD_FOLDER, g_name))
            db.execute("""
                INSERT INTO clothing_images (clothing_id, image, created_at)
                VALUES (?, ?, ?)
            """, (
                item_id,
                g_name,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))
    db.commit()

    # Generate Barcode (Code128) & QR Code
    # Barcode content: item id + name
    try:
        barcode_obj = Code128(str(item_id), writer=ImageWriter())
        barcode_filename_no_ext = f"barcode_{item_id}"
        barcode_full_path = os.path.join(UPLOAD_FOLDER, barcode_filename_no_ext)
        barcode_obj.save(barcode_full_path)  # creates .png
        barcode_filename = f"{barcode_filename_no_ext}.png"
    except Exception as e:
        barcode_filename = None

    try:
        qr_filename = f"qr_{item_id}.png"
        qr_data = f"ITEM:{item_id}|NAME:{name}|CAT:{category}"
        img_qr = qrcode.make(qr_data)
        img_qr.save(os.path.join(UPLOAD_FOLDER, qr_filename))
    except Exception as e:
        qr_filename = None

    db.execute(
        "UPDATE clothing SET barcode=?, qrcode=? WHERE id=?",
        (barcode_filename, qr_filename, item_id)
    )
    db.commit()

    # stock log
    db.execute("""
        INSERT INTO stock_logs (clothing_id, change_type, qty_change, note, admin, created_at)
        VALUES (?, 'create', ?, 'Initial stock', ?, ?)
    """, (
        item_id,
        quantity,
        session.get("admin", "system"),
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))
    db.commit()

    log_action("inventory_add", f"Added item '{name}' ({category}) x{quantity}")
    flash(f"✅ Item '{name}' added successfully!", "success")
    return redirect(url_for("inventory"))


# ---------------- EDIT ITEM (FORM) ----------------
@app.route("/inventory/edit/<int:item_id>")
@require_login
def edit_item(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM clothing WHERE id=?", (item_id,)).fetchone()
    if not item:
        return redirect(url_for("inventory"))

    categories = db.execute(
        "SELECT * FROM categories ORDER BY name ASC"
    ).fetchall()

    return render_template(
        "inventory_edit.html",
        item=item,
        categories=categories,
        title="Edit Item"
    )


# ---------------- UPDATE ITEM (POST) ----------------
@app.route("/inventory/update/<int:item_id>", methods=["POST"])
@require_login
def update_item(item_id):
    name = request.form["name"]
    category = request.form["category"]
    size = request.form["size"]
    quantity = int(request.form["quantity"])
    price = float(request.form["price"])

    image_file = request.files.get("image")

    db = get_db()
    item = db.execute("SELECT * FROM clothing WHERE id=?", (item_id,)).fetchone()

    old_qty = item["quantity"]
    old_image = item["image"]
    new_image = old_image

    # If new image uploaded -> delete old + save new
    if image_file and image_file.filename:
        if old_image:
            old_path = os.path.join(UPLOAD_FOLDER, old_image)
            if os.path.exists(old_path):
                os.remove(old_path)

        new_image = f"item_{item_id}_{image_file.filename}"
        image_file.save(os.path.join(UPLOAD_FOLDER, new_image))

    db.execute("""
        UPDATE clothing
        SET name=?, category=?, size=?, quantity=?, price=?, image=?
        WHERE id=?
    """, (name, category, size, quantity, price, new_image, item_id))
    db.commit()

    # stock change log if qty changed
    diff = quantity - old_qty
    if diff != 0:
        change_type = "in" if diff > 0 else "out"
        db.execute("""
            INSERT INTO stock_logs (clothing_id, change_type, qty_change, note, admin, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            item_id,
            change_type,
            diff,
            "Qty updated via Edit Item",
            session.get("admin", "system"),
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        db.commit()

    log_action("inventory_update", f"Updated item #{item_id} -> {name}")
    flash(f"✅ Item '{name}' updated successfully!", "success")
    return redirect(url_for("inventory"))


# ---------------- STOCK ADJUST PAGE ----------------
@app.route("/inventory/stock/<int:item_id>", methods=["GET", "POST"])
@require_login
def stock_adjust(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM clothing WHERE id=?", (item_id,)).fetchone()
    if not item:
        return redirect(url_for("inventory"))

    if request.method == "POST":
        change_type = request.form["change_type"]  # in / out / adjust
        qty = int(request.form["qty"])
        note = request.form.get("note", "").strip() or "-"

        current = item["quantity"]
        diff = qty

        if change_type == "in":
            new_qty = current + qty
            diff = qty
        elif change_type == "out":
            new_qty = max(0, current - qty)
            diff = new_qty - current   # negative or zero
        else:   # adjust => set to qty
            new_qty = qty
            diff = new_qty - current

        db.execute(
            "UPDATE clothing SET quantity=? WHERE id=?",
            (new_qty, item_id)
        )
        db.execute("""
            INSERT INTO stock_logs (clothing_id, change_type, qty_change, note, admin, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            item_id,
            change_type,
            diff,
            note,
            session.get("admin", "system"),
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        db.commit()

        log_action("stock_adjust", f"Item #{item_id} stock {change_type} ({diff})")
        flash(f"✅ Stock adjusted successfully! {change_type.title()} {abs(diff)} unit(s).", "success")
        return redirect(url_for("inventory"))

    return render_template(
        "stock_adjust.html",
        item=item,
        title="Adjust Stock"
    )


# ---------------- DELETE ITEM ----------------
@app.route("/inventory/delete/<int:item_id>")
@require_login
@require_role("admin", "superadmin")
def delete_item(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM clothing WHERE id=?", (item_id,)).fetchone()

    if item:
        # delete main image
        if item["image"]:
            img_path = os.path.join(UPLOAD_FOLDER, item["image"])
            if os.path.exists(img_path):
                os.remove(img_path)

        # delete codes
        if item["barcode"]:
            b_path = os.path.join(UPLOAD_FOLDER, item["barcode"])
            if os.path.exists(b_path):
                os.remove(b_path)
        if item["qrcode"]:
            q_path = os.path.join(UPLOAD_FOLDER, item["qrcode"])
            if os.path.exists(q_path):
                os.remove(q_path)

        # delete gallery images
        images = db.execute(
            "SELECT * FROM clothing_images WHERE clothing_id=?", (item_id,)
        ).fetchall()
        for img in images:
            p = os.path.join(UPLOAD_FOLDER, img["image"])
            if os.path.exists(p):
                os.remove(p)
        db.execute("DELETE FROM clothing_images WHERE clothing_id=?", (item_id,))

        db.execute("DELETE FROM clothing WHERE id=?", (item_id,))
        db.commit()

        log_action("inventory_delete", f"Deleted item '{item['name']}' (id={item_id})")
        flash(f"✅ Item '{item['name']}' deleted successfully!", "success")

    return redirect(url_for("inventory"))


# ---------------- GALLERY PAGES ----------------
@app.route("/inventory/gallery/<int:item_id>")
@require_login
def item_gallery(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM clothing WHERE id=?", (item_id,)).fetchone()
    if not item:
        return redirect(url_for("inventory"))

    imgs = db.execute(
        "SELECT * FROM clothing_images WHERE clothing_id=? ORDER BY id ASC",
        (item_id,)
    ).fetchall()

    return render_template(
        "gallery.html",
        item=item,
        images=imgs,
        title="Product Gallery"
    )


@app.route("/inventory/gallery/<int:item_id>/add", methods=["POST"])
@require_login
def item_gallery_add(item_id):
    db = get_db()
    files = request.files.getlist("gallery")
    for f in files:
        if f and f.filename:
            fname = f"gallery_{item_id}_{f.filename}"
            f.save(os.path.join(UPLOAD_FOLDER, fname))
            db.execute("""
                INSERT INTO clothing_images (clothing_id, image, created_at)
                VALUES (?, ?, ?)
            """, (
                item_id,
                fname,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))
    db.commit()
    flash(f"✅ {len(files)} image(s) added to gallery successfully!", "success")
    return redirect(url_for("item_gallery", item_id=item_id))


@app.route("/inventory/gallery/delete/<int:img_id>")
@require_login
def item_gallery_delete(img_id):
    db = get_db()
    row = db.execute("SELECT * FROM clothing_images WHERE id=?", (img_id,)).fetchone()
    if not row:
        return redirect(url_for("inventory"))
    item_id = row["clothing_id"]

    path = os.path.join(UPLOAD_FOLDER, row["image"])
    if os.path.exists(path):
        os.remove(path)

    db.execute("DELETE FROM clothing_images WHERE id=?", (img_id,))
    db.commit()
    flash(f"✅ Image deleted successfully!", "success")

    return redirect(url_for("item_gallery", item_id=item_id))


# ---------------- STOCK LOGS PAGE ----------------
@app.route("/stock-logs")
@require_login
def stock_logs():
    db = get_db()
    rows = db.execute("""
        SELECT s.*, c.name, c.category
        FROM stock_logs s
        LEFT JOIN clothing c ON c.id = s.clothing_id
        ORDER BY s.id DESC
        LIMIT 300
    """).fetchall()

    return render_template(
        "stock_logs.html",
        logs=rows,
        title="Stock Logs"
    )


# ---------------- EXPORT TO EXCEL ----------------
@app.route("/inventory/export")
@require_login
def export_inventory():
    db = get_db()
    rows = db.execute("""
        SELECT name, category, size, quantity, price, created_at
        FROM clothing
        ORDER BY name ASC
    """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["Name", "Category", "Size", "Quantity", "Price", "Created At"]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["name"], r["category"], r["size"],
            r["quantity"], r["price"], r["created_at"]
        ])

    export_name = "inventory_export.xlsx"
    export_path = os.path.join(UPLOAD_FOLDER, export_name)
    wb.save(export_path)

    return send_from_directory(
        UPLOAD_FOLDER,
        export_name,
        as_attachment=True
    )


# ---------------- IMPORT FROM EXCEL ----------------
@app.route("/inventory/import", methods=["POST"])
@require_login
@require_role("admin", "superadmin")
def import_inventory():
    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Please select an Excel file.", "danger")
        return redirect(url_for("inventory"))

    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    db = get_db()
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue  # skip header

        name, category, size, qty, price, created_at = row
        if not name:
            continue

        qty = int(qty or 0)
        price = float(price or 0)

        if not created_at:
            created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        db.execute("""
            INSERT INTO clothing (name, category, size, quantity, price, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (name, category, size, qty, price, str(created_at)))

        # stock log
        item_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        db.execute("""
            INSERT INTO stock_logs (clothing_id, change_type, qty_change, note, admin, created_at)
            VALUES (?, 'import', ?, 'Imported from Excel', ?, ?)
        """, (
            item_id,
            qty,
            session.get("admin", "system"),
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

    db.commit()
    log_action("inventory_import", f"Imported inventory from {file.filename}")
    flash(f"✅ Successfully imported {len(list(ws.iter_rows()))-1} items from Excel!", "success")

    return redirect(url_for("inventory"))

# ---------------- Regenerate Barcode/QR Routes ----------------
@app.route("/codes/regenerate/<int:item_id>")
@require_login
def regenerate_codes(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM clothing WHERE id=?", (item_id,)).fetchone()

    if not item:
        return redirect(url_for("inventory"))

    text = f"{item_id}-{item['name']}-{item['category']}"

    barcode_file = generate_barcode(item_id, text)
    qr_file = generate_qr(item_id, text)

    db.execute("UPDATE clothing SET barcode=?, qrcode=? WHERE id=?",
               (barcode_file, qr_file, item_id))
    db.commit()

    log_action("codes_regenerate", f"Regenerated codes for item #{item_id}")
    flash(f"✅ Barcode and QR code regenerated successfully!", "success")

    return redirect(url_for("inventory"))

# ---------------- Download ZIP of Codes ----------------
@app.route("/codes/zip/<int:item_id>")
def codes_zip(item_id):
    zipname = f"codes_{item_id}.zip"
    zippath = os.path.join(UPLOAD_FOLDER, zipname)

    zipf = zipfile.ZipFile(zippath, 'w', zipfile.ZIP_DEFLATED)
    zipf.write(os.path.join(UPLOAD_FOLDER, f"barcode_{item_id}.png"),
               f"barcode_{item_id}.png")
    zipf.write(os.path.join(UPLOAD_FOLDER, f"qr_{item_id}.png"),
               f"qr_{item_id}.png")
    zipf.close()

    return send_file(zippath, as_attachment=True)

# ---------------- Print Label Sheet (PDF) ----------------
@app.route("/codes/print/<int:item_id>")
def print_labels(item_id):
    pdfname = f"labels_{item_id}.pdf"
    pdfpath = os.path.join(UPLOAD_FOLDER, pdfname)

    c = canvas.Canvas(pdfpath, pagesize=A4)
    w, h = A4

    barcode_path = os.path.join(UPLOAD_FOLDER, f"barcode_{item_id}.png")
    qr_path = os.path.join(UPLOAD_FOLDER, f"qr_{item_id}.png")

    rows, cols = 10, 3
    x_gap, y_gap = 190, 80

    y = h - 80
    for r in range(rows):
        x = 40
        for col in range(cols):
            c.drawImage(barcode_path, x, y, width=150, height=40)
            c.drawImage(qr_path, x + 150, y, width=50, height=50)
            c.drawString(x, y - 10, f"Product #{item_id}")
            x += x_gap
        y -= y_gap

    c.save()
    return send_file(pdfpath, as_attachment=True)


# ---------------- CATEGORIES ----------------
# ---------------- CATEGORIES ----------------
@app.route("/categories")
@require_login
def categories():
    db = get_db()

    # search
    q = request.args.get("q", "").strip()

    # sorting
    sort = request.args.get("sort", "name_desc")
    sort_map = {
        "name_asc": "name ASC",
        "name_desc": "name DESC"
    }
    order_by = sort_map.get(sort, "name DESC")

    # pagination
    try:
        per_page = int(request.args.get("per_page", 10))
    except:
        per_page = 10

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    where = ""
    params = []

    if q:
        where = "WHERE name LIKE ? OR description LIKE ?"
        like = f"%{q}%"
        params.extend([like, like])

    # count
    count = db.execute(
        f"SELECT COUNT(*) AS c FROM categories {where}", params
    ).fetchone()["c"]

    total_pages = max(1, (count + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    cats = db.execute(
        f"""
        SELECT * FROM categories
        {where}
        ORDER BY {order_by}
        LIMIT ? OFFSET ?
        """,
        params + [per_page, offset]
    ).fetchall()

    pages = list(range(1, total_pages + 1))

    return render_template(
        "categories.html",
        cats=cats,
        q=q,
        sort=sort,
        per_page=per_page,
        page=page,
        pages=pages,
        count=count,
        title="Categories"
    )

@app.route("/categories/add", methods=["POST"])
@require_login
@require_role("admin", "superadmin")
def add_category():
    name = request.form["name"]
    desc = request.form.get("description", "")

    db = get_db()
    db.execute("INSERT OR IGNORE INTO categories (name, description) VALUES (?, ?)",
               (name, desc))
    db.commit()

    log_action("category_add", f"Added category '{name}'")
    flash(f"✅ Category '{name}' added successfully!", "success")
    return redirect(url_for("categories"))


@app.route("/categories/edit/<int:cat_id>", methods=["POST"])
@require_login
@require_role("admin", "superadmin")
def edit_category(cat_id):
    name = request.form["name"]
    desc = request.form.get("description", "")

    db = get_db()
    db.execute("UPDATE categories SET name=?, description=? WHERE id=?",
               (name, desc, cat_id))
    db.commit()

    log_action("category_edit", f"Edited category #{cat_id} -> {name}")
    flash(f"✅ Category '{name}' updated successfully!", "success")
    return redirect(url_for("categories"))


@app.route("/categories/delete/<int:cat_id>")
@require_login
@require_role("admin", "superadmin")
def delete_category(cat_id):
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE id=?", (cat_id,)).fetchone()
    db.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    db.commit()

    if cat:
        log_action("category_delete", f"Deleted category '{cat['name']}'")
        flash(f"✅ Category '{cat['name']}' deleted successfully!", "success")

    return redirect(url_for("categories"))


# ---------------- ACTIVITY LOGS ----------------
@app.route("/logs")
@require_login
@require_role("admin", "superadmin")
def logs():
    db = get_db()
    rows = db.execute("SELECT * FROM logs ORDER BY id DESC LIMIT 50").fetchall()
    return render_template("logs.html", logs=rows, title="Activity Logs")


# ---------------- SERVE UPLOADED IMAGES ----------------
@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


# ---------------- START APP ----------------
if __name__ == "__main__":
    init_db()
    app.run(debug=True)
